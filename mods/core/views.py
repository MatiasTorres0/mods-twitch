from django.http import request
from django.shortcuts import render, redirect
from django.contrib.auth import login as auth_login, authenticate # Import login as auth_login globally
from django.contrib import messages  # Asegúrate de importar messages
from django.views.generic import CreateView
from django.urls import reverse_lazy
from .forms import ComandoForm, ModeradorRegistroForm, AnunciosForm, Notas_ModsForms, Stream_WWEForms, Combate_WWEForms
from .models import Moderador, Comando, Notas_Mods, Stream_WWE, Combate_WWE # Asegúrate de que Comando esté importado

from django.contrib import messages
from django.db import transaction
from openpyxl import load_workbook
from django.contrib.auth.views import PasswordResetView, PasswordResetDoneView, PasswordResetConfirmView, PasswordResetCompleteView
from openpyxl import load_workbook
from rest_framework import viewsets
from .serializers import ComandoSerializer, ModeradorSerializer, NotasSerializer, Combate_WWESerializer



class ModeradorRegistroView(CreateView):
    model = Moderador
    form_class = ModeradorRegistroForm
    template_name = 'core/registro.html'
    success_url = reverse_lazy('login')
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Filtrar nombres de Twitch ya utilizados
        nombres_usados = Moderador.objects.values_list('nombre_twitch', flat=True)
        opciones_disponibles = [(key, value) for key, value in Moderador.MODERADOR if key not in nombres_usados]
        form.fields['nombre_twitch'].choices = [('', 'Selecciona tu nombre de Twitch')] + opciones_disponibles
        return form
    
    def form_valid(self, form):
        # Guardar el usuario primero
        user = form.save()
        
        try:
            # Autenticar y loguear al usuario después del registro
            username = form.cleaned_data.get('nombre_twitch')
            password = form.cleaned_data.get('password1')
            
            # Usar authenticate con el nombre de usuario correcto
            # Cambio aquí: usar nombre_twitch en lugar de username
            user_auth = authenticate(self.request, nombre_twitch=username, password=password)
            
            if user_auth is not None:
                auth_login(self.request, user_auth)  # Usar auth_login
            else:
                # Si la autenticación falla, registrar el error pero continuar
                print(f"Error de autenticación para el usuario {username}")
        except Exception as e:
            # Capturar cualquier excepción durante la autenticación
            print(f"Error durante la autenticación: {str(e)}")
        
        # Siempre redirigir al usuario, incluso si la autenticación falla
        return super().form_valid(form)

# Create your views here.
def index(request):
    return render(request, 'core/index.html')

def about(request):
    return render(request, 'core/about.html')

def contact(request):
    return render(request, 'core/contact.html')

def service(request):
    return render(request, 'core/service.html')

def blog(request):
    return render(request, 'core/blog.html')

def blog_details(request):
    return render(request, 'core/blog_details.html')

def registro(request):
    if request.method == 'POST':
        form = ModeradorRegistroForm(request.POST)
        print("Formulario enviado")
        print("Errores:", form.errors)
        if form.is_valid():
            user_obj = form.save()
            nombre_twitch = form.cleaned_data.get('nombre_twitch')
            password = form.cleaned_data.get('password1')
            
            # Autenticar usando nombre_twitch
            user_auth = authenticate(request, nombre_twitch=nombre_twitch, password=password)
            
            if user_auth is not None:
                auth_login(request, user_auth)
                messages.success(request, f'¡Registro exitoso! Recuerda que tu nombre de usuario para iniciar sesión es: {nombre_twitch}')
                return redirect('login')
            else:
                messages.error(request, "No se pudo autenticar después del registro.")
                return redirect('login')
    else:
        form = ModeradorRegistroForm()
    return render(request, 'registro/registro.html', {'form': form})

def login_view(request):
    if request.method == 'POST':
        nombre_twitch = request.POST.get('nombre_twitch')  # Cambiado de 'username'
        password = request.POST.get('password')
        
        user = authenticate(request, nombre_twitch=nombre_twitch, password=password)
        
        if user is None:
            try:
                posible_moderador = Moderador.objects.filter(nombre_twitch__iexact=nombre_twitch).first()
                if posible_moderador:
                    messages.error(request, 'Contraseña incorrecta. Por favor, intenta de nuevo.')
                else:
                    messages.error(request, f'No existe un moderador con el nombre "{nombre_twitch}". Verifica que sea tu nombre de Twitch exacto.')
            except Exception as e:
                print(f"Error during login: {str(e)}")
                messages.error(request, 'Nombre de Twitch o contraseña incorrectos.')
            return render(request, 'registro/login.html', {'error_message': 'Credenciales incorrectas'})
        
        auth_login(request, user)
        return redirect('inicio')
    
    return render(request, 'registro/login.html')

def inicio(request):
    if request.GET.get('ajax'):
        # Si es una solicitud AJAX, devolver solo el contenido
        return render(request, 'dashboard/inicio_content.html')
    # Si no es AJAX, devolver la página completa
    return render(request, 'dashboard/inicio.html')

def comandos(request):
    comandos = Comando.objects.all()
    data = {
        'comandos': comandos
    }
    return render(request, 'dashboard/comandos.html', data)

def protocolos(request):
    if request.GET.get('ajax'):
        return render(request, 'dashboard/protocolos_content.html')
    return render(request, 'dashboard/protocolos.html')

def anuncios(request):
    if request.GET.get('ajax'):
        return render(request, 'dashboard/anuncios_content.html')
    return render(request, 'dashboard/anuncios.html')

def agregar_comando(request):
    data = {
        'form': ComandoForm()
    }
    if request.method == 'POST':
        comando = ComandoForm(request.POST)
        if comando.is_valid():
            comando.save()
            data['mensaje'] = "El comando ha sido agregado con éxito"

            return redirect('comandos')
    else:
        comando = ComandoForm()
    
    context = {'comando': comando}
    
    
    return render(request, 'dashboard/agregar_comando.html', data)

def editar_comando(request, comando_id):
    # Obtener el comando a editar
    comando = Comando.objects.get(id=comando_id)
    data = {
        'form': ComandoForm(instance=comando)
    }

    if request.method == 'POST':
        form = ComandoForm(data=request.POST, instance = comando)
        if form.is_valid():
            form.save()
            data['mensaje'] = "El comando ha sido modificado con éxito"
            data['form'] = form
            return redirect('comandos')
        else:
            data['form'] = form
            data['mensaje'] = "Error al modificar el comando"
            
    return render(request, 'dashboard/editar_comando.html', data)

def editar_nota(request, nota_id):
    # Obtener la nota a editar
    nota = Notas_Mods.objects.get(id=nota_id)
    data = {
        'form': Notas_ModsForms(instance=nota)
    }

    if request.method == 'POST':
        form = Notas_ModsForms(data=request.POST, instance = nota)
        if form.is_valid():
            form.save()
            data['mensaje'] = "La nota ha sido modificada con éxito"
            data['form'] = form
            return redirect('ver_notas') # Assuming 'ver_notas' is the correct URL name to redirect to after editing
        else:
            data['form'] = form
            data['mensaje'] = "Error al modificar la nota"
            
    return render(request, 'dashboard/editar_nota.html', data)

def eliminar_nota(request, nota_id):
    # Obtener el comando a eliminar
    nota = Notas_Mods.objects.get(id=nota_id)

    # Eliminar el comando
    nota.delete()

    # Redirigir a la página de comandos o a donde desees
    return redirect('ver_notas')


def eliminar_comando(request, comando_id):
    # Obtener el comando a eliminar
    comando = Comando.objects.get(id=comando_id)

    # Eliminar el comando
    comando.delete()

    # Redirigir a la página de comandos o a donde desees
    return redirect('comandos')

def protocolos(request):
    return render(request, 'dashboard/protocolos.html')

def anuncios(request):
    return render(request, 'dashboard/anuncios.html')

def logout_view(request):
    from django.contrib.auth import logout
    logout(request)
    messages.success(request, 'Has cerrado sesión correctamente.')
    return redirect('index')


class CustomPasswordResetView(PasswordResetView):
    template_name = 'core/password_reset_form.html'
    email_template_name = 'core/password_reset_email.html'
    success_url = reverse_lazy('password_reset_done')

class CustomPasswordResetDoneView(PasswordResetDoneView):
    template_name = 'core/password_reset_done.html'

class CustomPasswordResetConfirmView(PasswordResetConfirmView):
    template_name = 'core/password_reset_confirm.html'
    success_url = reverse_lazy('password_reset_complete')

class CustomPasswordResetCompleteView(PasswordResetCompleteView):
    template_name = 'core/password_reset_complete.html'


def upload_excel_comandos(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, 'Formato de archivo no válido')
            return redirect('comandos')

        try:
            wb = load_workbook(excel_file)
            sheet = wb.active
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) < 3:  # Cambiado a 3 columnas
                    messages.warning(request, f'Fila inválida: {row}')
                    continue
                
                # Eliminado el bucle for row in rows redundante
                Comando.objects.create(
                    nombre_comando=row[0],
                    tipo_comando=row[1],  # Segunda columna como categoría
                    descripcion=row[2],   # Tercera columna como descripción
                    ejemplo=row[2],       # Usar la descripción como ejemplo
                    activo=True
                )
                
            messages.success(request, f'{sheet.max_row -1} comandos importados!')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
        
        return redirect('comandos')
    
    return render(request, 'dashboard/agregar_comandos_excel.html')


class ComandoViewSet(viewsets.ModelViewSet):
    queryset = Comando.objects.all()
    serializer_class = ComandoSerializer

class ModeradorViewSet(viewsets.ModelViewSet):
    queryset = Moderador.objects.all()
    serializer_class = ModeradorSerializer

def agregar_anuncio(request):
    if request.method == 'POST':
        form = AnunciosForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('anuncios')
    else:
        form = AnunciosForm()
    return render(request, 'dashboard/agregar_anuncio.html', {'form': form})

def politicas_uso_app(request):
    return render(request, 'core/politicas_uso_app.html')

def notas_mods(request):
    if request.method == 'POST':
        form = Notas_ModsForms(request.POST)
        if form.is_valid():
            form.save()
            return redirect('notas_mods')
    else:
        form = Notas_ModsForms()
    return render(request, 'dashboard/notas_mods.html', {'form': form})

def agregar_combate(request):
    # Initialize forms
    form_stream = Stream_WWEForms()
    form_combate = Combate_WWEForms()
    
    if request.method == 'POST':
        if 'submit_stream' in request.POST:
            form_stream = Stream_WWEForms(request.POST)
            if form_stream.is_valid():
                form_stream.save()
                messages.success(request, 'Stream WWE agregado correctamente')
                return redirect('agregar_combate')
            else:
                messages.error(request, 'Error al agregar el Stream WWE')
                
        elif 'submit_combate' in request.POST:
            form_combate = Combate_WWEForms(request.POST)
            if form_combate.is_valid():
                form_combate.save()
                messages.success(request, 'Combate WWE agregado correctamente')
                return redirect('agregar_combate')
            else:
                messages.error(request, 'Error al agregar el Combate WWE')
    
    # Create context with both forms
    context = {
        'form_stream': form_stream,
        'form_combate': form_combate
    }
    
    return render(request, 'dashboard/agregar_combate.html', context)
def seccion_agregar(request):
    return render(request, 'dashboard/seccion_agregar.html')

def ver_notas(request):
    notas = Notas_Mods.objects.all()
    return render(request, 'dashboard/ver_notas.html', {'notas': notas})

def subir_excel_combate(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        fecha_stream = request.POST.get('fecha_stream')
        
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, 'Formato de archivo no válido')
            return redirect('agregar_combate')
        
        if not fecha_stream:
            messages.error(request, 'Debe especificar la fecha del stream')
            return redirect('agregar_combate')

        try:
            # Crear el stream principal
            with transaction.atomic():
                stream_instance = Stream_WWE.objects.create(
                    fecha_stream=fecha_stream
                )
                
                wb = load_workbook(excel_file)
                sheet = wb.active
                
                for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                    # Saltar filas vacías o encabezados
                    if not row[0] or row[0] == "Categoría/Evento":
                        continue
                    
                    categoria = row[0]
                    
                    # Determinar tipo de combate
                    if "Parejas" in categoria:
                        tipo_combate = 'PAREJAS'
                    elif "Especial" in categoria:
                        tipo_combate = 'ESPECIAL'
                    else:
                        tipo_combate = 'INDIVIDUAL'
                    
                    # Limpiar y obtener participantes válidos
                    participantes = []
                    for cell in row[1:]:
                        if cell and str(cell).strip() and str(cell).strip().lower() not in ['vs', 'y']:
                            participantes.append(str(cell).strip())
                    
                    # Crear instancia de combate
                    combate = Combate_WWE(
                        stream=stream_instance,
                        tipo_combate=tipo_combate,
                        resultado='NO_CONTEST',
                        orden_combate=idx
                    )
                    
                    # Asignar participantes según el tipo
                    if tipo_combate == 'INDIVIDUAL' and len(participantes) >= 2:
                        combate.luchador_1 = participantes[0]
                        combate.luchador_2 = participantes[1]
                    elif tipo_combate == 'PAREJAS' and len(participantes) >= 4:
                        combate.luchador_1 = participantes[0]
                        combate.luchador_2 = participantes[1]
                        combate.luchador_3 = participantes[2]
                        combate.luchador_4 = participantes[3]
                    elif tipo_combate == 'ESPECIAL' and participantes:
                        combate.luchador_1 = participantes[0]
                    else:
                        continue  # Saltar filas inválidas
                    
                    combate.save()
                
                messages.success(request, f'¡Stream y {idx-1} combates importados correctamente!')
        
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
        
        return redirect('agregar_combate')
    
    return render(request, 'dashboard/agregar_combate_excel.html')

def ver_combates(request):
    combates = Combate_WWE.objects.all()
    return render(request, 'dashboard/ver_combates.html', {'combates': combates})


class NotasViewSet(viewsets.ModelViewSet):
    queryset = Notas_Mods.objects.all()
    serializer_class = NotasSerializer

class CombateViewSet(viewsets.ModelViewSet):
    queryset = Combate_WWE.objects.all()
    serializer_class = Combate_WWESerializer